import openpyxl
import json
import os

os.makedirs("data", exist_ok=True)

print("Converting property tax data...")
wb = openpyxl.load_workbook("property_tax_2024.xlsx", data_only=True)
ws = wb["Master"]
headers = []
for cell in next(ws.iter_rows(min_row=1, max_row=1)):
    headers.append(cell.value)
rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if any(c is not None for c in row):
        obj = {}
        for i, val in enumerate(row):
            key = headers[i]
            if isinstance(val, float):
                val = round(val, 4)
            obj[key] = val
        rows.append(obj)
with open("data/property_tax.json", "w") as f:
    json.dump(rows, f)
print("Done: " + str(len(rows)) + " records saved to data/property_tax.json")

print("Converting benchmarking data...")
wb2 = openpyxl.load_workbook("benchmarking_master_2024.xlsx", data_only=True)
ws2 = wb2["master"]
headers2 = []
for cell in next(ws2.iter_rows(min_row=1, max_row=1)):
    headers2.append(cell.value)
rows2 = []
for row in ws2.iter_rows(min_row=2, values_only=True):
    if any(c is not None for c in row):
        obj = {}
        for i, val in enumerate(row):
            key = headers2[i]
            if isinstance(val, float):
                val = round(val, 4)
            obj[key] = val
        rows2.append(obj)
with open("data/benchmarking.json", "w") as f:
    json.dump(rows2, f)
print("Done: " + str(len(rows2)) + " records saved to data/benchmarking.json")

print("Converting local government data...")
wb3 = openpyxl.load_workbook("local_government_2024.xlsx", data_only=True)
all_local = []
for sheet in ["County", "City", "Town", "Village"]:
    ws3 = wb3[sheet]
    headers3 = []
    for cell in next(ws3.iter_rows(min_row=1, max_row=1)):
        headers3.append(cell.value)
    for row in ws3.iter_rows(min_row=2, values_only=True):
        if any(c is not None for c in row):
            obj = {}
            for i, val in enumerate(row):
                key = headers3[i]
                if isinstance(val, float):
                    val = round(val, 4)
                obj[key] = val
            obj["entity_type"] = sheet.lower()
            all_local.append(obj)
with open("data/local_gov.json", "w") as f:
    json.dump(all_local, f)
print("Done: " + str(len(all_local)) + " records saved to data/local_gov.json")

print("All done! Check your data folder.")
